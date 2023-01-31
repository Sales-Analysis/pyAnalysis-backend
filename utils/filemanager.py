import csv
import openpyxl

from code_errors import FileIsEmptyError, HeaderNotFoundError
from typing import Dict, Union, List
from models import InputModel, ABCModels


def read_exel(path: str) -> InputModel:
    """Читает входной exel файл и возвращает в виде словаря."""
    wb = openpyxl.load_workbook(path)
    worksheet = wb.worksheets[0]
    iter_rows = worksheet.iter_rows()

    try:
        header: List[str] = [cell.value for cell in next(iter_rows)]
    except StopIteration:
        raise FileIsEmptyError
    data = convert_dict(header=header, rows=iter_rows)
    if not len([value for values in data.values() for value in values]):
        raise FileIsEmptyError
    result = InputModel(
        CODE_PLU=data[ABCModels.CODE_PLU],
        NAME_ANALYSIS_POSITIONS=data[ABCModels.NAME_ANALYSIS_POSITIONS],
        DATA_ANALYSIS=data[ABCModels.DATA_ANALYSIS],
    )
    return result


def convert_dict(
    header: List[str],
    rows: List[Dict[str, List[Union[int, str, float]]]]
) -> Dict[str, List[Union[int, str, float]]]:
    """Конвертирует в один словарь."""
    result: Dict[str, List[Union[int, str, float]]] = {name: [] for name in header}
    for row in rows:
        for cell, name in zip(row, header):
            result[name].append(cell.value)
    return result


def read_csv(path: str) -> Dict[str, List[Union[int, str, float]]]:
    result: List[List[Union[float, int, str]]] = []
    header: List[str] = []
    with open(path, encoding='UTF-8') as f:
        reader = csv.reader(f, delimiter=';')
        for index, row in enumerate(reader):
            if index == 0:
                if not row:
                    raise HeaderNotFoundError
                header = row
            elif not row:
                continue
            else:
                result.append(row)
    if not result:
        raise FileIsEmptyError
    data = {i: [] for i in header}
    for i, v in enumerate(result):
        for j, k in enumerate(v):
            data[header[j]].append(k)
    return data
